"""Parse uploaded Excel workbooks into BIAP's normalized business dataset shape."""
from __future__ import annotations

import base64
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def parse_excel_base64(payload: str, *, name: str = "Excel data") -> dict[str, Any]:
    try:
        raw = base64.b64decode(payload, validate=True)
    except Exception as exc:
        raise ValueError("invalid Excel base64 payload") from exc
    if not raw:
        raise ValueError("Excel file is empty")
    if len(raw) > 12 * 1024 * 1024:
        raise ValueError("Excel file is too large")

    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("unable to open Excel workbook") from exc

    if not workbook.sheetnames:
        raise ValueError("Excel workbook has no sheets")
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel sheet must contain a header and at least one data row")

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(header):
        raise ValueError("Excel header row is empty")
    columns: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(header):
        base = value or f"column_{index + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        columns.append(base if count == 0 else f"{base}_{count + 1}")

    normalized: list[dict[str, str]] = []
    for values in rows[1:50001]:
        if values is None:
            continue
        record = {column: "" if i >= len(values) or values[i] is None else str(values[i]) for i, column in enumerate(columns)}
        if any(value.strip() for value in record.values()):
            normalized.append(record)
    if not normalized:
        raise ValueError("Excel sheet contains no usable data rows")

    return {
        "name": name.strip() or "Excel data",
        "columns": columns,
        "rows": normalized,
        "source": "xlsx-file",
    }
